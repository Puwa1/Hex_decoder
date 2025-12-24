import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import tkinter.font as tkfont
import pandas as pd
import os
import datetime
from decoder import UniversalJT808Decoder

# --- MATPLOTLIB IMPORTS ---
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates

class ModernApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GPS Decoder Pro (Modern UI + Analytics)")
        self.root.geometry("1450x900")
        self.root.configure(bg="#f8fafc") 

        # --- STYLE CONFIGURATION ---
        self.style = ttk.Style()
        self.style.theme_use('clam') 

        # 1. Colors Palette
        BG_COLOR = "#f8fafc"
        WHITE    = "#ffffff"
        PRIMARY  = "#3b82f6"    # Blue
        PRIMARY_HOVER = "#2563eb"
        SUCCESS  = "#10b981"    # Green
        SUCCESS_HOVER = "#059669"
        WARNING  = "#f59e0b"    # Orange
        WARNING_HOVER = "#d97706"
        DANGER   = "#ef4444"    # Red
        DANGER_HOVER  = "#dc2626"
        NEUTRAL  = "#64748b"    # Slate
        NEUTRAL_HOVER = "#475569"
        INFO     = "#8b5cf6"    # Violet
        INFO_HOVER = "#7c3aed"
        
        TEXT_COLOR = "#334155" 
        HEADER_BG  = "#e2e8f0"

        self.colors = {
            "primary": PRIMARY, "primary_hover": PRIMARY_HOVER,
            "success": SUCCESS, "success_hover": SUCCESS_HOVER,
            "warning": WARNING, "warning_hover": WARNING_HOVER,
            "danger": DANGER,   "danger_hover": DANGER_HOVER,
            "neutral": NEUTRAL, "neutral_hover": NEUTRAL_HOVER,
            "info":    INFO,    "info_hover":    INFO_HOVER
        }

        # 2. General Settings
        self.style.configure(".", background=BG_COLOR, foreground=TEXT_COLOR, font=('Segoe UI', 10))
        self.style.configure("TFrame", background=BG_COLOR)
        self.style.configure("TLabelframe", background=BG_COLOR, borderwidth=1, relief='flat')
        self.style.configure("TLabelframe.Label", background=BG_COLOR, foreground="#475569", font=('Segoe UI', 10, 'bold'))

        # 3. Treeview Styling
        self.style.configure("Treeview", 
            background=WHITE, fieldbackground=WHITE, foreground="#334155", 
            rowheight=32, borderwidth=0, font=('Segoe UI', 10)
        )
        self.style.configure("Treeview.Heading", 
            background=HEADER_BG, foreground="#1e293b", 
            font=('Segoe UI', 10, 'bold'), relief="flat"
        )
        self.style.map('Treeview', 
            background=[('selected', '#dbeafe')], 
            foreground=[('selected', '#1e40af')]
        )

        # 4. Buttons Styling
        self.btn_configs = {
            "Success.TButton": (SUCCESS, SUCCESS_HOVER),
            "Primary.TButton": (PRIMARY, PRIMARY_HOVER),
            "Warning.TButton": (WARNING, WARNING_HOVER),
            "Danger.TButton":  (DANGER, DANGER_HOVER),
            "Neutral.TButton": (NEUTRAL, NEUTRAL_HOVER),
            "Info.TButton":    (INFO, INFO_HOVER)
        }
        for style_name, (bg, active_bg) in self.btn_configs.items():
            self.style.configure(style_name, background=bg, foreground=WHITE, font=('Segoe UI', 10, 'bold'), borderwidth=0, focuscolor='none')
            self.style.map(style_name, background=[('active', active_bg)])
        
        # 5. Notebook
        self.style.configure("TNotebook", background=BG_COLOR, borderwidth=0)
        self.style.configure("TNotebook.Tab", 
            background="#cbd5e1", foreground="#64748b",
            padding=[20, 10], font=('Segoe UI', 10, 'bold'), borderwidth=0
        )
        self.style.map("TNotebook.Tab", 
            background=[("selected", WHITE)], 
            foreground=[("selected", PRIMARY)]
        )

        self.config_file = 'master_mapping_config.csv'
        self.decoder = UniversalJT808Decoder(self.config_file)
        self.results_df = None
        self.config_df = None
        self.last_loaded_paths = None 
        self.font_measure = tkfont.Font(family='Segoe UI', size=10)
        
        # --- LAYOUT ---
        main_container = ttk.Frame(root)
        main_container.pack(fill='both', expand=True, padx=20, pady=20)

        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(expand=True, fill='both')
        
        self.tab_0200 = ttk.Frame(self.notebook); self.notebook.add(self.tab_0200, text="📍 Single Trace")
        self.tab_0704 = ttk.Frame(self.notebook); self.notebook.add(self.tab_0704, text="📦 Batch Trace")
        self.tab_file = ttk.Frame(self.notebook); self.notebook.add(self.tab_file, text="📂 File Process")
        self.tab_graph = ttk.Frame(self.notebook); self.notebook.add(self.tab_graph, text="📊 Analytics") 
        self.tab_config = ttk.Frame(self.notebook); self.notebook.add(self.tab_config, text="⚙️ Settings") 
        
        self.setup_ui()
        self.setup_graph_ui() 
        self.setup_config_editor() 

        if self.decoder.config is not None: self.log(f"System Ready: Loaded {self.config_file}", "success")
        else: self.log("⚠️ Warning: Configuration file missing.", "error")

    def setup_ui(self):
        log_f = ttk.LabelFrame(self.root, text=" System Log ")
        log_f.pack(side='bottom', fill='x', padx=20, pady=(0, 20))
        
        self.txt_log = scrolledtext.ScrolledText(log_f, height=5, font=('Consolas', 10), bg="#1e293b", fg="#e2e8f0", borderwidth=0)
        self.txt_log.pack(fill='both', padx=5, pady=5)
        self.txt_log.tag_config('success', foreground='#4ade80') 
        self.txt_log.tag_config('error', foreground='#f87171')   
        self.txt_log.tag_config('warning', foreground='#fcd34d')

        self.setup_tab(self.tab_0200, "0200", "Success.TButton")
        self.setup_tab(self.tab_0704, "0704", "Warning.TButton")
        
        # File Tab Layout
        f_container = ttk.Frame(self.tab_file)
        f_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        self.lbl_protocol = ttk.Label(f_container, text=f"Current Protocol: {os.path.basename(self.config_file)}", foreground="#64748b", font=('Segoe UI', 9, 'italic'))
        self.lbl_protocol.pack(fill='x', pady=(0, 5))

        btn_bar = ttk.Frame(f_container)
        btn_bar.pack(fill='x', pady=(0, 15))
        
        tk.Button(btn_bar, text="⚙️ Change Protocol", command=self.change_protocol_file, 
                 bg=self.colors['info'], fg="white", font=('Segoe UI', 10, 'bold'), relief='flat', padx=15, pady=8, cursor='hand2').pack(side='left', padx=(0, 10))

        tk.Button(btn_bar, text="📂 Load CSV Data", command=self.select_file, 
                 bg=self.colors['primary'], fg="white", font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=8, cursor='hand2').pack(side='left', padx=(0, 10))
                 
        tk.Button(btn_bar, text="🔄 Refresh Data", command=self.refresh_data, 
                 bg=self.colors['neutral'], fg="white", font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=8, cursor='hand2').pack(side='left', padx=(0, 10))

        tk.Button(btn_bar, text="💾 Export Report", command=self.export, 
                 bg=self.colors['success'], fg="white", font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=8, cursor='hand2').pack(side='left')

        self.create_table(f_container, "file")

    def setup_tab(self, frame, mode, btn_style):
        container = ttk.Frame(frame)
        container.pack(fill='both', expand=True, padx=20, pady=20)

        top_bar = ttk.Frame(container)
        top_bar.pack(fill='x', pady=(0, 15))
        
        lbl = ttk.Label(top_bar, text="Paste Hex String:", font=('Segoe UI', 11, 'bold'))
        lbl.pack(anchor='w', pady=(0, 5))

        entry = scrolledtext.ScrolledText(top_bar, height=3, font=('Consolas', 11), borderwidth=1, relief="solid")
        entry.pack(fill='x', pady=(0, 10))
        
        if mode == "0200": self.txt_0200 = entry
        else: self.txt_0704 = entry
        
        ttk.Button(top_bar, text=f"⚡ DECODE {mode}", command=lambda: self.run(mode), style=btn_style, cursor='hand2').pack(fill='x', ipady=5)

        self.create_table(container, mode)

    def create_table(self, frame, name):
        table_frame = ttk.Frame(frame)
        table_frame.pack(expand=True, fill='both')

        tree = ttk.Treeview(table_frame, show='headings')
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        tree.bind("<Double-1>", lambda e: self.show_row_details(e, tree))
        if os.name == 'nt': tree.bind("<Button-3>", lambda event: self.show_column_menu(event, tree))
        else: tree.bind("<Button-2>", lambda event: self.show_column_menu(event, tree))
        
        tree.tag_configure('error_row', background='#fee2e2', foreground='#b91c1c') 
        tree.tag_configure('even_row', background='#f8fafc') 
        
        if name == "0200": self.tree_0200 = tree
        elif name == "0704": self.tree_0704 = tree
        else: self.tree_file = tree

    # ==========================
    # --- GRAPH ANALYTICS TAB ---
    # ==========================
    def setup_graph_ui(self):
        container = ttk.Frame(self.tab_graph)
        container.pack(fill='both', expand=True, padx=20, pady=20)

        # Control Bar
        ctrl_frame = ttk.Frame(container)
        ctrl_frame.pack(fill='x', pady=(0, 15))

        ttk.Label(ctrl_frame, text="Select Graph Type:", font=('Segoe UI', 11, 'bold')).pack(side='left', padx=(0,10))
        
        self.graph_type = tk.StringVar(value="Speed Profile")
        # [UPDATED] Changed 'Alarm Summary' to 'Alarm Sign'
        options = ["Speed Profile", "Fuel Level", "Alarm Sign", "Route Scatter"]
        
        cb = ttk.Combobox(ctrl_frame, textvariable=self.graph_type, values=options, state="readonly", width=25, font=('Segoe UI', 10))
        cb.pack(side='left', padx=(0, 10))
        cb.bind("<<ComboboxSelected>>", self.plot_graph)

        ttk.Button(ctrl_frame, text="📊 Plot Graph", command=self.plot_graph, style="Primary.TButton").pack(side='left')

        # Canvas Area
        self.graph_frame = ttk.Frame(container, style="TFrame")
        self.graph_frame.pack(fill='both', expand=True)
        self.canvas_widget = None

    def plot_graph(self, event=None):
        if self.results_df is None or self.results_df.empty:
            messagebox.showwarning("No Data", "Please load and decode data first in 'File Process' tab.")
            return

        df = self.results_df.copy()
        time_col = next((c for c in df.columns if 'time' in c.lower()), None)
        if time_col:
            df[time_col] = pd.to_datetime(df[time_col], errors='coerce')
            df = df.sort_values(by=time_col)
        
        for widget in self.graph_frame.winfo_children(): widget.destroy()

        fig, ax = plt.subplots(figsize=(10, 5), dpi=100)
        mode = self.graph_type.get()

        try:
            if mode == "Speed Profile":
                speed_col = next((c for c in df.columns if 'speed' in c.lower()), None)
                if speed_col and time_col:
                    df[speed_col] = pd.to_numeric(df[speed_col], errors='coerce')
                    ax.plot(df[time_col], df[speed_col], color='#3b82f6', linewidth=2)
                    ax.fill_between(df[time_col], df[speed_col], color='#3b82f6', alpha=0.1)
                    ax.set_title("Vehicle Speed Profile", fontsize=12, fontweight='bold')
                    ax.set_ylabel("Speed (km/h)")
                    ax.set_xlabel("Time")
                    ax.grid(True, linestyle='--', alpha=0.5)
                    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
                else: ax.text(0.5, 0.5, "Missing Speed or Time column", ha='center', va='center')

            elif mode == "Fuel Level":
                fuel_col = next((c for c in df.columns if 'fuel' in c.lower()), None)
                if fuel_col and time_col:
                    df[fuel_col] = pd.to_numeric(df[fuel_col], errors='coerce')
                    ax.plot(df[time_col], df[fuel_col], color='#10b981', linewidth=2)
                    ax.set_title("Fuel Level Monitoring", fontsize=12, fontweight='bold')
                    ax.set_ylabel("Fuel Level")
                    ax.set_xlabel("Time")
                    ax.grid(True, linestyle='--', alpha=0.5)
                    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
                else: ax.text(0.5, 0.5, "Missing Fuel or Time column", ha='center', va='center')

            # [UPDATED] Changed condition to match new option
            elif mode == "Alarm Sign":
                # พยายามหา Column ที่มีคำว่า 'alarm' หรือ 'sign'
                alarm_col = next((c for c in df.columns if 'alarm' in c.lower() or 'sign' in c.lower()), None)
                if alarm_col:
                    counts = df[alarm_col].value_counts()
                    counts = counts[counts.index != 'Normal'] # Filter Normal
                    if not counts.empty:
                        wedges, texts, autotexts = ax.pie(counts, labels=counts.index, autopct='%1.1f%%', startangle=90, colors=plt.cm.Pastel1.colors)
                        ax.set_title("Alarm Sign Distribution", fontsize=12, fontweight='bold')
                    else: ax.text(0.5, 0.5, "No Alarms Found (All Normal)", ha='center', va='center')
                else: ax.text(0.5, 0.5, "Missing 'Alarm/Sign' column", ha='center', va='center')

            elif mode == "Route Scatter":
                lat_col = next((c for c in df.columns if 'lat' in c.lower()), None)
                lon_col = next((c for c in df.columns if 'lon' in c.lower() or 'lng' in c.lower()), None)
                if lat_col and lon_col:
                    df[lat_col] = pd.to_numeric(df[lat_col], errors='coerce')
                    df[lon_col] = pd.to_numeric(df[lon_col], errors='coerce')
                    sc = ax.scatter(df[lon_col], df[lat_col], c=df.index, cmap='viridis', s=10, alpha=0.7)
                    ax.set_title("Route Scatter Plot (Lat/Lon)", fontsize=12, fontweight='bold')
                    ax.set_ylabel("Latitude")
                    ax.set_xlabel("Longitude")
                    ax.grid(True, linestyle='--', alpha=0.5)
                else: ax.text(0.5, 0.5, "Missing Latitude/Longitude columns", ha='center', va='center')

        except Exception as e:
            ax.text(0.5, 0.5, f"Error plotting graph: {str(e)}", ha='center', va='center', color='red')

        canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    # ==========================
    # --- CONFIG EDITOR ---
    # ==========================
    def setup_config_editor(self):
        container = ttk.Frame(self.tab_config)
        container.pack(fill='both', expand=True, padx=20, pady=20)

        toolbar = ttk.Frame(container)
        toolbar.pack(fill='x', pady=(0, 10))
        
        ttk.Button(toolbar, text="🔄 Reload File", command=self.load_config_data, style="Neutral.TButton").pack(side='left', padx=(0, 5))
        ttk.Button(toolbar, text="➕ Add Row (Insert)", command=self.add_config_row_inline, style="Primary.TButton").pack(side='left', padx=(0, 5))
        ttk.Button(toolbar, text="❌ Delete Row", command=self.delete_config_row, style="Danger.TButton").pack(side='left', padx=(0, 5))
        
        ttk.Button(toolbar, text="💾 Save System", command=self.save_config_data, style="Success.TButton").pack(side='right', padx=(5, 0))
        ttk.Button(toolbar, text="📥 Import Backup", command=self.import_config_backup, style="Warning.TButton").pack(side='right', padx=(5, 0))
        ttk.Button(toolbar, text="📤 Export Backup", command=self.export_config_backup, style="Neutral.TButton").pack(side='right')

        columns = ['Category', 'SubID', 'FieldName', 'StartByte', 'Length', 'DataType', 'Scale']
        self.tree_config = ttk.Treeview(container, columns=columns, show='headings', selectmode='browse')
        
        vsb = ttk.Scrollbar(container, orient="vertical", command=self.tree_config.yview)
        self.tree_config.configure(yscrollcommand=vsb.set)
        
        self.tree_config.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        for col in columns:
            self.tree_config.heading(col, text=col)
            self.tree_config.column(col, width=100, anchor='center')
        
        self.tree_config.column('FieldName', width=200, anchor='w')
        self.tree_config.bind("<Double-1>", self.on_tree_double_click)
        self.load_config_data()

    def load_config_data(self):
        try:
            if os.path.exists(self.config_file):
                self.config_df = pd.read_csv(self.config_file, dtype=str)
                self.config_df.fillna('', inplace=True)
                self.tree_config.delete(*self.tree_config.get_children())
                for i, row in self.config_df.iterrows():
                    vals = [row[c] for c in self.tree_config['columns']]
                    tags = ('even_row',) if i % 2 == 0 else ()
                    self.tree_config.insert("", "end", values=vals, tags=tags)
                self.log(f"Editor Loaded: {os.path.basename(self.config_file)}", "success")
            else:
                self.log("Config file not found.", "error")
        except Exception as e:
            self.log(f"Error loading config: {e}", "error")

    def change_protocol_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("CSV Protocol", "*.csv")],
            title="Select Protocol Configuration File"
        )
        if not path: return
        self.config_file = path
        self.decoder = UniversalJT808Decoder(self.config_file)
        self.lbl_protocol.config(text=f"Current Protocol: {os.path.basename(self.config_file)}")
        self.load_config_data()
        self.refresh_data(silent=True)
        self.log(f"Protocol Switched to: {os.path.basename(path)}", "success")
        messagebox.showinfo("Protocol Changed", f"Active protocol is now:\n{os.path.basename(path)}")

    def export_config_backup(self):
        try:
            rows = []
            columns = self.tree_config['columns']
            for item in self.tree_config.get_children():
                rows.append(self.tree_config.item(item)['values'])
            df_to_save = pd.DataFrame(rows, columns=columns)

            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV File", "*.csv")], title="Export Configuration Backup")
            if not path: return
            df_to_save.to_csv(path, index=False)
            self.log(f"Config exported to {os.path.basename(path)}", "success")
            messagebox.showinfo("Export Success", f"Configuration backed up successfully!")
        except Exception as e:
            self.log(f"Export failed: {e}", "error")
            messagebox.showerror("Export Error", str(e))

    def import_config_backup(self):
        try:
            path = filedialog.askopenfilename(filetypes=[("CSV File", "*.csv")], title="Import Configuration Backup")
            if not path: return
            new_df = pd.read_csv(path, dtype=str)
            new_df.fillna('', inplace=True)
            self.tree_config.delete(*self.tree_config.get_children())
            for i, row in new_df.iterrows():
                vals = [row[c] for c in self.tree_config['columns']]
                tags = ('even_row',) if i % 2 == 0 else ()
                self.tree_config.insert("", "end", values=vals, tags=tags)
            new_df.to_csv(self.config_file, index=False)
            self.decoder = UniversalJT808Decoder(self.config_file)
            self.refresh_data(silent=True)
            self.log(f"Imported rules from {os.path.basename(path)}", "success")
            messagebox.showinfo("Import Success", "Configuration imported and system updated!")
        except Exception as e:
            self.log(f"Import failed: {e}", "error")
            messagebox.showerror("Import Error", f"Invalid Config File: {e}")

    def on_tree_double_click(self, event):
        region = self.tree_config.identify("region", event.x, event.y)
        if region != "cell": return
        column = self.tree_config.identify_column(event.x)
        row_id = self.tree_config.identify_row(event.y)
        if not row_id: return
        col_idx = int(column.replace('#', '')) - 1
        x, y, w, h = self.tree_config.bbox(row_id, column)
        current_val = self.tree_config.item(row_id, 'values')[col_idx]
        entry = ttk.Entry(self.tree_config)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, str(current_val))
        entry.focus()
        entry.select_range(0, tk.END)
        def save_edit(event_inner=None):
            new_val = entry.get()
            current_values = list(self.tree_config.item(row_id, 'values'))
            current_values[col_idx] = new_val
            self.tree_config.item(row_id, values=current_values)
            entry.destroy()
        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit) 
        entry.bind("<Escape>", lambda e: entry.destroy())

    def add_config_row_inline(self):
        default_vals = ['BSJ_Ext', '0000', 'New Field', '0', '4', 'DWORD', '1']
        selected = self.tree_config.selection()
        if selected:
            target_id = selected[-1]
            target_index = self.tree_config.index(target_id)
            new_item = self.tree_config.insert("", target_index + 1, values=default_vals)
        else:
            new_item = self.tree_config.insert("", "end", values=default_vals)
        self.tree_config.selection_set(new_item)
        self.tree_config.focus(new_item)
        self.tree_config.see(new_item)

    def delete_config_row(self):
        selected = self.tree_config.selection()
        if not selected: return
        if messagebox.askyesno("Confirm Delete", "Delete selected row(s)?"):
            for item in selected:
                self.tree_config.delete(item)

    def save_config_data(self):
        try:
            rows = []
            columns = self.tree_config['columns']
            for item in self.tree_config.get_children():
                rows.append(self.tree_config.item(item)['values'])
            new_df = pd.DataFrame(rows, columns=columns)
            new_df.to_csv(self.config_file, index=False)
            self.decoder = UniversalJT808Decoder(self.config_file)
            self.log("✅ Config saved and System updated.", "success")
            self.refresh_data(silent=True)
            if not getattr(self, '_silent_save', False):
                 messagebox.showinfo("Success", "Saved! Data has been refreshed.")
        except Exception as e:
            self.log(f"Error saving config: {e}", "error")
            messagebox.showerror("Error", f"Failed to save: {e}")

    def show_column_menu(self, event, tree):
        menu = tk.Menu(self.root, tearoff=0, bg="white", fg="black")
        all_cols = list(tree['columns'])
        visible_cols = tree['displaycolumns']
        if visible_cols == '#all': visible_cols = all_cols
        else: visible_cols = list(visible_cols)
        menu.add_command(label="👁️ Show All Columns", command=lambda: self.reset_columns(tree))
        menu.add_separator()
        for col in all_cols:
            is_visible = col in visible_cols
            menu.add_checkbutton(label=col, onvalue=1, offvalue=0, variable=tk.IntVar(value=1 if is_visible else 0), command=lambda c=col, t=tree: self.toggle_column(t, c))
        menu.post(event.x_root, event.y_root)

    def toggle_column(self, tree, col_name):
        all_cols = list(tree['columns'])
        current_display = list(tree['displaycolumns'])
        if current_display == ['#all']: current_display = all_cols
        if col_name in current_display: current_display.remove(col_name)
        else:
            current_display.append(col_name)
            current_display.sort(key=lambda x: all_cols.index(x) if x in all_cols else 999)
        if len(current_display) == 0: return 
        tree['displaycolumns'] = current_display

    def reset_columns(self, tree): tree['displaycolumns'] = '#all'

    def autosize_columns(self, tree, df):
        if df is None: df = pd.DataFrame()
        all_config_cols = self.decoder.all_field_names 
        extra_cols = [c for c in df.columns if c not in all_config_cols]
        final_cols = all_config_cols + extra_cols
        tree['columns'] = final_cols
        tree['displaycolumns'] = '#all' 
        for col in final_cols:
            tree.heading(col, text=col, anchor='center')
            max_width = 150 
            header_w = self.font_measure.measure(str(col)) + 40 
            if header_w > max_width: max_width = header_w
            if col in df.columns and len(df) > 0:
                sample_val = str(df[col].iloc[0])
                w = self.font_measure.measure(sample_val) + 40
                if w > max_width: max_width = w
            if max_width > 500: max_width = 500
            tree.column(col, width=max_width, minwidth=100, anchor='center', stretch=False)
        return final_cols

    def log(self, message, level="info"):
        self.txt_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n", level)
        self.txt_log.see(tk.END)

    def run(self, mode):
        raw = self.txt_0200.get("1.0", tk.END).strip() if mode=="0200" else self.txt_0704.get("1.0", tk.END).strip()
        if not raw: return
        decoded, st, _ = self.decoder.decode_raw(raw)
        self.display(decoded, mode)

    def display(self, data, mode):
        tree = self.tree_0200 if mode=="0200" else self.tree_0704
        tree.delete(*tree.get_children())
        if not data: self.log("No data found.", "error"); return
        if data[0].get('_Status', '').startswith('❌'):
            self.log(f"Decode Failed: {data[0]['_Status']}", "error")
        else:
            self.log(f"Successfully decoded {len(data)} records.", "success")
        df = pd.DataFrame(data)
        final_cols = self.autosize_columns(tree, df)
        for i, (_, r) in enumerate(df.iterrows()):
            tags = []
            if not str(r.get('_Status', 'OK')).startswith('OK'):
                tags.append('error_row')
            elif i % 2 == 0:
                tags.append('even_row')
            tree.insert("", "end", values=[r.get(c,"") for c in final_cols], tags=tags)

    def select_file(self):
        paths = filedialog.askopenfilenames()
        if not paths: return
        self.last_loaded_paths = paths 
        self.process_batch_files(paths)

    def refresh_data(self, silent=False):
        self.decoder = UniversalJT808Decoder(self.config_file)
        if self.last_loaded_paths:
            self.process_batch_files(self.last_loaded_paths)
            if not silent: self.log("🔄 Data Refreshed.", "success")
        else:
            if not silent: self.log("⚠️ No file loaded to refresh.", "warning")

    def process_batch_files(self, paths):
        res = []
        for p in paths:
            try:
                df = pd.read_csv(p)
                col = next((c for c in df.columns if c.lower() in ['raw-data','hex','raw']), None)
                if col:
                    for _, r in df.iterrows():
                        d, _, _ = self.decoder.decode_raw(str(r[col]))
                        res.extend(d)
            except: pass
        self.results_df = pd.DataFrame(res)
        self.tree_file.delete(*self.tree_file.get_children())
        if not res: return
        final_cols = self.autosize_columns(self.tree_file, self.results_df)
        for i, (_, r) in enumerate(self.results_df.iterrows()):
            tags = []
            if not str(r.get('_Status', 'OK')).startswith('OK'):
                tags.append('error_row')
            elif i % 2 == 0:
                tags.append('even_row')
            self.tree_file.insert("", "end", values=[r.get(c,"") for c in final_cols], tags=tags)
        self.log(f"Batch Processed: {len(res)} rows loaded.", "success")

    def export(self):
        if self.results_df is None: return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel File", "*.xlsx"), ("CSV File", "*.csv")])
        if not file_path: return
        try:
            export_df = self.results_df.copy()
            non_empty = [c for c in export_df.columns if export_df[c].astype(str).str.strip().ne('').any()]
            keep = [c for c in self.decoder.all_field_names if c in non_empty]
            keep.extend([c for c in non_empty if c not in keep])
            export_df = export_df[keep]
            if file_path.lower().endswith('.csv'):
                export_df.to_csv(file_path, index=False, encoding='utf-8-sig')
                self.log(f"Saved CSV: {file_path}", "success")
            else:
                try:
                    export_df.to_excel(file_path, index=False)
                    import openpyxl
                    from openpyxl.styles import PatternFill, Font
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    status_col_idx = None
                    for idx, cell in enumerate(ws[1], 1):
                        if cell.value == '_Status': status_col_idx = idx; break
                    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                    header_font = Font(bold=True)
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    if status_col_idx:
                        for row in ws.iter_rows(min_row=2):
                            status_val = str(row[status_col_idx-1].value)
                            if not status_val.startswith('OK'):
                                for cell in row: cell.fill = red_fill
                    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 22
                    wb.save(file_path)
                    self.log(f"Saved Excel (Colored): {file_path}", "success")
                except Exception as e_excel:
                    export_df.to_excel(file_path, index=False)
                    self.log(f"Saved Excel (Plain): {file_path}", "success")
        except Exception as e: self.log(f"Export Error: {e}", "error")