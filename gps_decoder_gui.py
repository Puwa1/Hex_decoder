import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import tkinter.font as tkfont
import pandas as pd
import binascii
import threading
import os
import datetime

# --- PART 1: LOGIC (เหมือนเดิม 100%) ---
class UniversalJT808Decoder:
    def __init__(self, config_file='master_mapping_config.csv'):
        self.config = None
        self.rules_map = {} 
        self.all_field_names = [] 
        
        self.status_definitions = {
            0: ("ACC:OFF", "ACC:ON"), 1: ("GPS:NoFix", "GPS:Fixed"),
            2: ("Lat:North", "Lat:South"), 3: ("Lng:East", "Lng:West"),
            10: ("Oil:Normal", "Oil:Disconnect"), 4: ("Op:Normal", "Op:Stop"), 5: ("Enc:No", "Enc:Yes")
        }
        self.alarm_definitions = {
            0: "SOS", 1: "OverSpeed", 7: "MainPower:LowVolt", 8: "MainPower:Off", 29: "Collision",
            2: "Fatigue", 3: "EarlyWarn", 4: "GNSS:Fault", 5: "GNSS:Cut", 6: "Power:Low",
            9: "LCD:Cut", 18: "Drive:Time", 19: "Stop:Time", 20: "Area:In/Out"
        }
        self.dtc_status_map = {
            0x1E: "Stop Idle", 0x1F: "Running", 0x20: "Running", 0x21: "Running",
            0x22: "Change to Idle", 0x32: "Period Idle", 0x37: "Card In", 0x38: "Card Out"
        }

        if os.path.exists(config_file):
            try:
                self.config = pd.read_csv(config_file)
                self.config['SubID'] = self.config['SubID'].fillna('').astype(str).str.strip().str.upper()
                for _, row in self.config.iterrows():
                    key = f"{row['Category']}{row['SubID']}"
                    self.rules_map[key] = row.to_dict()

                raw_names = self.config['FieldName'].unique().tolist()
                
                self.all_field_names = [
                    'Time', 'Message ID', 'Body Attr Raw', 'Body Length',
                    'Protocol Version', 'Terminal ID', 'Count Up Track',
                    'Alarm Flags', 'Alarm Desc', 'Terminal Status', 'Status Desc',
                    'Latitude', 'Longitude', 'Altitude (m)', 'Speed (km/h)', 'Course (deg)',
                    'Mileage (km)', 'Fuel Level (L)', 'Speed Recorder (km/h)', 'Video Alarm', 
                    'Storage Fault', 'Signal Status', 'GSM Signal', 'Satellites',
                    'Alarm Sleep Status', 'Lock Sim Status', 'Base Station (LBS)', 
                    'Ext Voltage (V)', 'Device IMEI',
                    'Mobile Operator', 'DTC Status', 'R-Value', 'Special Value', 
                    'GPS Status', 'Backup Battery (V)', 'Buffer Remaining', 'HDOP',
                    'Firmware Ver', 'Driver ID Info', 'BSJ Serial No',
                    'Raw Hex Block', '_Status'
                ]
            except Exception as e: print(f"Error reading CSV: {e}")

    def unescape(self, content_hex):
        return content_hex.replace('7D02', '7E').replace('7D01', '7D')

    def verify_checksum(self, data_bytes):
        try:
            if len(data_bytes) < 2: return False
            calc = 0
            for b in data_bytes[:-1]: calc ^= b
            return (calc == data_bytes[-1])
        except: return False

    def format_timestamp(self, bcd):
        if not bcd or len(bcd) != 12: return bcd
        return f"20{bcd[0:2]}-{bcd[2:4]}-{bcd[4:6]} {bcd[6:8]}:{bcd[8:10]}:{bcd[10:12]}"

    def parse_value(self, val_hex, rule):
        try:
            dtype = rule['DataType']; scale = float(rule['Scale']) if rule['Scale'] else 1.0
            if dtype in ['WORD', 'DWORD', 'BYTE']: return int(val_hex, 16) / scale
            elif dtype == 'SIGNED_WORD':
                raw = int(val_hex, 16); return (raw - 0x10000 if raw >= 0x8000 else raw) / scale
            elif dtype == 'SIGNED_DWORD':
                raw = int(val_hex, 16); return (raw - 0x100000000 if raw >= 0x80000000 else raw) / scale
            elif dtype == 'STRING':
                return binascii.unhexlify(val_hex).decode('gbk', errors='ignore').strip('\x00').strip()
            return val_hex
        except: return f"ERR:{val_hex}"

    def parse_bsj_tlv(self, container_hex):
        res = {}; idx = 0; limit = len(container_hex)
        while idx < limit:
            try:
                if idx+4 > limit: break
                length_val = int(container_hex[idx:idx+4], 16)
                if idx+8 > limit: break 
                sub_id = container_hex[idx+4:idx+8]
                
                data_len_hex = (length_val - 2) * 2
                data_start = idx+8
                if data_start + data_len_hex > limit: break
                val_hex = container_hex[data_start : data_start + data_len_hex]
                
                if sub_id == "0110" and len(val_hex) >= 34:
                    res['Mobile Operator'] = {0x44:'DTAC', 0x41:'AIS', 0x54:'TRUE'}.get(int(val_hex[0:2], 16), f"{val_hex[0:2]}")
                    res['DTC Status'] = self.dtc_status_map.get(int(val_hex[2:4], 16), str(int(val_hex[2:4], 16)))
                    res['R-Value'] = int(val_hex[4:12], 16)
                    res['Special Value'] = int(val_hex[12:20], 16)
                    res['GPS Status'] = bytes.fromhex(val_hex[20:22]).decode('ascii') if len(val_hex)>=22 else ""
                    res['Backup Battery (V)'] = int(val_hex[22:26], 16) / 10.0
                    res['Buffer Remaining'] = int(val_hex[26:30], 16)
                    res['HDOP'] = int(val_hex[30:34], 16) / 100.0

                key = f"BSJ_Ext{sub_id}"
                if key in self.rules_map:
                    res[self.rules_map[key]['FieldName']] = self.parse_value(val_hex, self.rules_map[key])
                idx += 4 + (length_val * 2) 
            except: break
        return res

    def decode_raw(self, raw_hex):
        try:
            error_msgs = []
            hex_str = str(raw_hex).upper().replace(" ", "").replace("\n", "").replace("\r", "").strip()
            
            if not (hex_str.startswith('7E') and hex_str.endswith('7E')):
                error_msgs.append("No 7E Header/Tail")

            try:
                content_hex = hex_str[2:-2] if (hex_str.startswith('7E') and hex_str.endswith('7E')) else hex_str
                data_str = self.unescape(content_hex)
                data_bytes = bytes.fromhex(data_str)
            except:
                error_msgs.append("Invalid Hex Chars")
                return [{'Message ID': 'HEX ERR', '_Status': " | ".join(error_msgs), 'Raw Hex Block': hex_str}], "Error", 0

            if not self.verify_checksum(data_bytes):
                error_msgs.append("Checksum Fail")

            if error_msgs:
                return [{'Message ID': f"0x{data_str[0:4]}" if len(data_str)>=4 else '?', '_Status': "❌ " + " | ".join(error_msgs), 'Raw Hex Block': hex_str}], "Error", 0

            row_status = "OK"
            data = data_str
            body_attr = int(data[4:8], 16)
            is_2019 = (body_attr >> 14) & 1
            header = {'Message ID': f"0x{data[0:4]}", 'Body Attr Raw': f"0x{body_attr:04X}", 'Body Length': body_attr & 0x03FF}
            
            if is_2019:
                header['Protocol Version'] = int(data[8:10], 16)
                header['Terminal ID'] = data[10:30]
                header['Count Up Track'] = int(data[30:34], 16)
                body_start = 34
            else:
                header['Protocol Version'] = "-"
                header['Terminal ID'] = data[8:20]
                header['Count Up Track'] = int(data[20:24], 16)
                body_start = 24
            if (body_attr >> 13) & 1: body_start += 8

            def parse_one_msg(b_hex):
                res = {}
                res['Alarm Flags'] = f"0x{b_hex[0:8]}"
                res['Terminal Status'] = f"0x{b_hex[8:16]}"
                st_val = int(b_hex[8:16], 16); al_val = int(b_hex[0:8], 16)
                res['Status Desc'] = f"{'ACC:ON' if st_val&1 else 'ACC:OFF'} | {'GPS:Fixed' if st_val&2 else 'GPS:NoFix'}"
                res['Alarm Desc'] = self.alarm_definitions.get(al_val, "Normal" if al_val==0 else str(al_val))
                res['Latitude'] = int(b_hex[16:24], 16) / 1000000.0
                res['Longitude'] = int(b_hex[24:32], 16) / 1000000.0
                res['Altitude (m)'] = int(b_hex[32:36], 16) if int(b_hex[32:36], 16) < 0x8000 else int(b_hex[32:36], 16) - 0x10000
                res['Speed (km/h)'] = int(b_hex[36:40], 16) / 10.0
                res['Course (deg)'] = int(b_hex[40:44], 16)
                res['Time'] = self.format_timestamp(b_hex[44:56])
                
                ext_hex = b_hex[56:]
                i = 0
                while i < len(ext_hex):
                    try:
                        if i+4 > len(ext_hex): break
                        eid = ext_hex[i:i+2]; e_len = int(ext_hex[i+2:i+4], 16)
                        if i+4+(e_len*2) > len(ext_hex): break
                        e_val = ext_hex[i+4 : i+4+(e_len*2)]
                        if eid in ['EB', 'EC', 'ED', 'EF']: res.update(self.parse_bsj_tlv(e_val))
                        else:
                            key = f"Extension{eid}"
                            if key in self.rules_map: res[self.rules_map[key]['FieldName']] = self.parse_value(e_val, self.rules_map[key])
                        i += 4 + (e_len * 2)
                    except: break
                return {**header, **res, 'Raw Hex Block': b_hex}

            decoded_list = []
            mid = int(data[0:4], 16); body_hex = data[body_start:-2]
            if mid == 0x0704:
                try:
                    count = int(body_hex[0:4], 16); curr = 6
                    for _ in range(count):
                        l_len = int(body_hex[curr:curr+4], 16)
                        decoded_list.append(parse_one_msg(body_hex[curr+4 : curr+4+(l_len*2)]))
                        curr += 4 + (l_len*2)
                except: decoded_list.append(parse_one_msg(body_hex))
            else: decoded_list.append(parse_one_msg(body_hex))

            for item in decoded_list: item['_Status'] = row_status
            return decoded_list, row_status, mid
        except Exception as e: 
            return [{'Message ID': 'CRASH', '_Status': f"❌ Sys Error: {str(e)}"}], "Error", 0

# --- PART 2: GUI (MODERN DESIGN + ZEBRA STRIPING) ---
class ModernApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GPS Decoder Pro (Modern UI)")
        self.root.geometry("1450x900")
        self.root.configure(bg="#f3f4f6") 

        # --- STYLE CONFIGURATION ---
        self.style = ttk.Style()
        self.style.theme_use('clam') 

        # 1. Colors Palette
        BG_COLOR = "#f3f4f6"
        WHITE = "#ffffff"
        PRIMARY = "#2563eb"    # Blue
        SUCCESS = "#16a34a"    # Green
        WARNING = "#ea580c"    # Orange
        TEXT_COLOR = "#1f2937" 
        HEADER_BG = "#e5e7eb"

        # 2. General Settings
        self.style.configure(".", background=BG_COLOR, foreground=TEXT_COLOR, font=('Segoe UI', 10))
        self.style.configure("TFrame", background=BG_COLOR)
        self.style.configure("TLabelframe", background=BG_COLOR, borderwidth=1, relief='flat')
        self.style.configure("TLabelframe.Label", background=BG_COLOR, foreground="#4b5563", font=('Segoe UI', 10, 'bold'))

        # 3. Treeview (Table) Styling
        self.style.configure("Treeview", 
            background=WHITE, 
            fieldbackground=WHITE, 
            foreground="#374151", 
            rowheight=35,
            borderwidth=0,
            font=('Segoe UI', 10)
        )
        self.style.configure("Treeview.Heading", 
            background=HEADER_BG, 
            foreground="#111827", 
            font=('Segoe UI', 10, 'bold'),
            relief="flat"
        )
        self.style.map('Treeview', background=[('selected', '#dbeafe')])

        # 4. Buttons Styling
        self.style.configure("Success.TButton", 
            background=SUCCESS, foreground=WHITE, 
            font=('Segoe UI', 10, 'bold'), borderwidth=0, focuscolor='none'
        )
        self.style.map("Success.TButton", background=[('active', '#15803d')])

        self.style.configure("Primary.TButton", 
            background=PRIMARY, foreground=WHITE, 
            font=('Segoe UI', 10, 'bold'), borderwidth=0, focuscolor='none'
        )
        self.style.map("Primary.TButton", background=[('active', '#1d4ed8')])

        self.style.configure("Warning.TButton", 
            background=WARNING, foreground=WHITE, 
            font=('Segoe UI', 10, 'bold'), borderwidth=0, focuscolor='none'
        )
        self.style.map("Warning.TButton", background=[('active', '#c2410c')])
        
        # 5. Notebook (Tabs)
        self.style.configure("TNotebook", background=BG_COLOR, borderwidth=0)
        self.style.configure("TNotebook.Tab", 
            background="#d1d5db", foreground="#4b5563", 
            padding=[15, 8], font=('Segoe UI', 10, 'bold'), borderwidth=0
        )
        self.style.map("TNotebook.Tab", 
            background=[("selected", WHITE)], 
            foreground=[("selected", PRIMARY)]
        )

        self.decoder = UniversalJT808Decoder('master_mapping_config.csv')
        self.results_df = None
        self.font_measure = tkfont.Font(family='Segoe UI', size=10)
        
        # --- LAYOUT ---
        main_container = ttk.Frame(root)
        main_container.pack(fill='both', expand=True, padx=20, pady=20)

        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(expand=True, fill='both')
        
        self.tab_0200 = ttk.Frame(self.notebook); self.notebook.add(self.tab_0200, text="📍 Single Trace (0200)")
        self.tab_0704 = ttk.Frame(self.notebook); self.notebook.add(self.tab_0704, text="📦 Batch Trace (0704)")
        self.tab_file = ttk.Frame(self.notebook); self.notebook.add(self.tab_file, text="📂 Batch File Processing")
        
        self.setup_ui()
        if self.decoder.config is not None: self.log(f"System Ready: Config loaded successfully.", "success")
        else: self.log("⚠️ Warning: Configuration file missing.", "error")

    def setup_ui(self):
        log_f = ttk.LabelFrame(self.root, text=" System Log ")
        log_f.pack(side='bottom', fill='x', padx=20, pady=(0, 20))
        
        self.txt_log = scrolledtext.ScrolledText(log_f, height=5, font=('Consolas', 10), bg="#1e293b", fg="#e2e8f0", borderwidth=0)
        self.txt_log.pack(fill='both', padx=5, pady=5)
        self.txt_log.tag_config('success', foreground='#4ade80') 
        self.txt_log.tag_config('error', foreground='#f87171')   

        self.setup_tab(self.tab_0200, "0200", "Success.TButton")
        self.setup_tab(self.tab_0704, "0704", "Warning.TButton")
        
        # File Tab Layout
        f_container = ttk.Frame(self.tab_file)
        f_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        btn_bar = ttk.Frame(f_container)
        btn_bar.pack(fill='x', pady=(0, 15))
        
        tk.Button(btn_bar, text="📂 Load CSV File", command=self.load_file, 
                 bg="#2563eb", fg="white", font=('Segoe UI', 11, 'bold'), relief='flat', padx=20, pady=8, cursor='hand2').pack(side='left', padx=(0, 10))
                 
        tk.Button(btn_bar, text="💾 Export Report", command=self.export, 
                 bg="#059669", fg="white", font=('Segoe UI', 11, 'bold'), relief='flat', padx=20, pady=8, cursor='hand2').pack(side='left')

        self.create_table(f_container, "file")

    def setup_tab(self, frame, mode, btn_style):
        container = ttk.Frame(frame)
        container.pack(fill='both', expand=True, padx=20, pady=20)

        top_bar = ttk.Frame(container)
        top_bar.pack(fill='x', pady=(0, 15))
        
        lbl = ttk.Label(top_bar, text="Paste Hex String:", font=('Segoe UI', 11, 'bold'))
        lbl.pack(anchor='w', pady=(0, 5))

        entry = scrolledtext.ScrolledText(top_bar, height=3, font=('Consolas', 11), borderwidth=1, relief="solid")
        entry.pack(fill='x', pady=(0, 10))
        
        if mode == "0200": self.txt_0200 = entry
        else: self.txt_0704 = entry
        
        ttk.Button(top_bar, text=f"⚡ DECODE {mode}", command=lambda: self.run(mode), style=btn_style, cursor='hand2').pack(fill='x', ipady=5)

        self.create_table(container, mode)

    def create_table(self, frame, name):
        table_frame = ttk.Frame(frame)
        table_frame.pack(expand=True, fill='both')

        tree = ttk.Treeview(table_frame, show='headings')
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        tree.bind("<Double-1>", lambda e: self.show_row_details(e, tree))
        if os.name == 'nt': tree.bind("<Button-3>", lambda event: self.show_column_menu(event, tree))
        else: tree.bind("<Button-2>", lambda event: self.show_column_menu(event, tree))
        
        # --- ZEBRA STRIPING TAGS ---
        # Error: Red
        tree.tag_configure('error_row', background='#fee2e2', foreground='#b91c1c') 
        # Even Rows: Subtle Gray
        tree.tag_configure('even_row', background='#f9fafb') 
        # Odd Rows: White (Default)
        
        if name == "0200": self.tree_0200 = tree
        elif name == "0704": self.tree_0704 = tree
        else: self.tree_file = tree

    def show_column_menu(self, event, tree):
        menu = tk.Menu(self.root, tearoff=0, bg="white", fg="black")
        all_cols = list(tree['columns'])
        visible_cols = tree['displaycolumns']
        if visible_cols == '#all': visible_cols = all_cols
        else: visible_cols = list(visible_cols)

        menu.add_command(label="👁️ Show All Columns", command=lambda: self.reset_columns(tree))
        menu.add_separator()

        for col in all_cols:
            is_visible = col in visible_cols
            menu.add_checkbutton(
                label=col, onvalue=1, offvalue=0, 
                variable=tk.IntVar(value=1 if is_visible else 0),
                command=lambda c=col, t=tree: self.toggle_column(t, c)
            )
        menu.post(event.x_root, event.y_root)

    def toggle_column(self, tree, col_name):
        all_cols = list(tree['columns'])
        current_display = list(tree['displaycolumns'])
        if current_display == ['#all']: current_display = all_cols
        if col_name in current_display: current_display.remove(col_name)
        else:
            current_display.append(col_name)
            current_display.sort(key=lambda x: all_cols.index(x) if x in all_cols else 999)
        if len(current_display) == 0: return 
        tree['displaycolumns'] = current_display

    def reset_columns(self, tree): tree['displaycolumns'] = '#all'

    def autosize_columns(self, tree, df):
        if df is None or df.empty: return
        non_empty = [c for c in df.columns if df[c].astype(str).str.strip().ne('').any()]
        keep = [c for c in self.decoder.all_field_names if c in non_empty]
        extra = [c for c in non_empty if c not in keep]
        final_cols = keep + extra

        tree['columns'] = final_cols
        tree['displaycolumns'] = '#all' 

        for col in final_cols:
            tree.heading(col, text=col, anchor='center')
            max_width = 150 
            header_w = self.font_measure.measure(str(col)) + 40 
            if header_w > max_width: max_width = header_w
            
            if col in df.columns and len(df) > 0:
                sample_val = str(df[col].iloc[0])
                w = self.font_measure.measure(sample_val) + 40
                if w > max_width: max_width = w
            
            if max_width > 500: max_width = 500
            tree.column(col, width=max_width, minwidth=100, anchor='center', stretch=False)
        return final_cols

    def log(self, message, level="info"):
        self.txt_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n", level)
        self.txt_log.see(tk.END)

    def run(self, mode):
        raw = self.txt_0200.get("1.0", tk.END).strip() if mode=="0200" else self.txt_0704.get("1.0", tk.END).strip()
        if not raw: return
        decoded, st, _ = self.decoder.decode_raw(raw)
        self.display(decoded, mode)

    def display(self, data, mode):
        tree = self.tree_0200 if mode=="0200" else self.tree_0704
        tree.delete(*tree.get_children())
        if not data: self.log("No data found.", "error"); return
        
        if data[0].get('_Status', '').startswith('❌'):
            self.log(f"Decode Failed: {data[0]['_Status']}", "error")
        else:
            self.log(f"Successfully decoded {len(data)} records.", "success")
        
        df = pd.DataFrame(data)
        final_cols = self.autosize_columns(tree, df)
        
        # --- APPLY ALTERNATING COLORS LOGIC ---
        for i, (_, r) in enumerate(df.iterrows()):
            tags = []
            # 1. Check Error First (Priority)
            if not str(r.get('_Status', 'OK')).startswith('OK'):
                tags.append('error_row')
            # 2. Else, check Even row for gray background
            elif i % 2 == 0:
                tags.append('even_row')
            
            tree.insert("", "end", values=[r.get(c,"") for c in final_cols], tags=tags)

    def load_file(self):
        paths = filedialog.askopenfilenames()
        if not paths: return
        res = []
        for p in paths:
            try:
                df = pd.read_csv(p)
                col = next((c for c in df.columns if c.lower() in ['raw-data','hex','raw']), None)
                if col:
                    for _, r in df.iterrows():
                        d, _, _ = self.decoder.decode_raw(str(r[col]))
                        res.extend(d)
            except: pass
        self.results_df = pd.DataFrame(res)
        self.tree_file.delete(*self.tree_file.get_children())
        if not res: return
        final_cols = self.autosize_columns(self.tree_file, self.results_df)
        
        # --- APPLY ALTERNATING COLORS LOGIC ---
        for i, (_, r) in enumerate(self.results_df.iterrows()):
            tags = []
            if not str(r.get('_Status', 'OK')).startswith('OK'):
                tags.append('error_row')
            elif i % 2 == 0:
                tags.append('even_row')
                
            self.tree_file.insert("", "end", values=[r.get(c,"") for c in final_cols], tags=tags)
            
        self.log(f"Batch Processed: {len(res)} rows loaded.", "success")

    def show_row_details(self, event, tree):
        item_id = tree.identify_row(event.y)
        if not item_id: return
        values = tree.item(item_id, 'values')
        cols = tree['columns']
        
        popup = tk.Toplevel(self.root)
        popup.title("📄 Row Details")
        popup.geometry("600x700")
        popup.configure(bg="#f3f4f6")
        
        txt = scrolledtext.ScrolledText(popup, font=('Consolas', 11), bg="white", fg="#333", borderwidth=0, padx=10, pady=10)
        txt.pack(fill='both', expand=True, padx=20, pady=20)
        
        for c, v in zip(cols, values): 
            txt.insert(tk.END, f"{c:<30}: {v}\n")
            txt.insert(tk.END, "-"*60 + "\n")
        
        txt.config(state='disabled')

    def export(self):
        if self.results_df is None: return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx"), ("CSV File", "*.csv")]
        )
        if not file_path: return
        
        try:
            export_df = self.results_df.copy()
            non_empty = [c for c in export_df.columns if export_df[c].astype(str).str.strip().ne('').any()]
            keep = [c for c in self.decoder.all_field_names if c in non_empty]
            keep.extend([c for c in non_empty if c not in keep])
            export_df = export_df[keep]

            if file_path.lower().endswith('.csv'):
                export_df.to_csv(file_path, index=False, encoding='utf-8-sig')
                self.log(f"Saved CSV: {file_path}", "success")
            else:
                try:
                    export_df.to_excel(file_path, index=False)
                    import openpyxl
                    from openpyxl.styles import PatternFill, Font, Border, Side
                    
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    
                    status_col_idx = None
                    for idx, cell in enumerate(ws[1], 1):
                        if cell.value == '_Status': status_col_idx = idx; break
                    
                    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                    header_font = Font(bold=True)
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    if status_col_idx:
                        for row in ws.iter_rows(min_row=2):
                            status_val = str(row[status_col_idx-1].value)
                            if not status_val.startswith('OK'):
                                for cell in row: cell.fill = red_fill
                                    
                    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 22
                        
                    wb.save(file_path)
                    self.log(f"Saved Excel (Colored): {file_path}", "success")
                except Exception as e_excel:
                    print(f"Coloring Error: {e_excel}")
                    export_df.to_excel(file_path, index=False)
                    self.log(f"Saved Excel (Plain): {file_path}", "success")
        except Exception as e: self.log(f"Export Error: {e}", "error")

if __name__ == "__main__":
    root = tk.Tk(); app = ModernApp(root); root.mainloop()